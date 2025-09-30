import os
from openai import OpenAI
import json
import pandas as pd
from datetime import date, datetime
import shutil
from datetime import datetime

# Helper to get OpenAI client with API key from env or Streamlit secrets
def get_openai_client():
    api_key = os.environ.get("OPENAI_API_KEY")
    if api_key:
        return OpenAI(api_key=api_key)
    try:
        import streamlit as st
        api_key = st.secrets["OPENAI_API_KEY"]
        return OpenAI(api_key=api_key)
    except Exception:
        return OpenAI()  # fallback, may fail if no key

def upload_image_for_vision(image_path: str) -> str:
    """
    Uploads an image to OpenAI Vision API and returns the image id.
    Args:
        image_path (str): Path to the image file.
    Returns:
        str: The image id assigned by OpenAI.
    """
    client = get_openai_client()
    with open(image_path, "rb") as file_content:
        image = client.files.create(
            file=file_content,
            purpose="vision",
        )
    print(f"Uploaded image with file_id: {image.id}")
    
    # Store file_id in param.json
    with open("param.json", "w", encoding="utf-8") as f:
        json.dump({"file_id": image.id}, f, indent=4)

    return image.id

def query_vision_model(image_id: str, json_file: str, gpt_model: str, prompt: str):
    """
    Queries the OpenAI Vision model with the given image_id and prompt, saves the result as golf_df.json, and returns True if the response is valid JSON and saves it to json_file.
    Args:
        image_id (str): The file_id of the uploaded image.
        gpt_model (str): The model name to use.
        prompt (str): The prompt for the Vision API.
    Returns:
        bool: True if the response is valid JSON and was saved, False otherwise.
    """
    client = get_openai_client()
    response = client.responses.create(
        model=gpt_model,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_image", "file_id": image_id},
            ],
        }],
    )
    print("Response ID:", response.id)
    try:
        data = json.loads(response.output_text)
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"Valid JSON response saved to {json_file}")
        return True
    except Exception as e:
        print(f"Invalid JSON response: {e}")
        print("Response text:", response.output_text)
        return False

def print_scorecard_table_from_json(json_path):
    """
    Loads a DataFrame from a JSON file and prints the scorecard table with fixed widths and a separator after the 10th column.
    Args:
        json_path (str): Path to the JSON file containing the DataFrame.
    """
    df = pd.DataFrame(json.load(open(json_path, "r", encoding="utf-8")))
    holes = list(df['hole'])
    pars = list(df['par'])
    hcps = list(df['hcp'])
    player_names = [col for col in df.columns if col not in ['hole', 'par', 'hcp']]

    def format_row(first, values):
        row = f"{str(first):<8}"
        for i, v in enumerate(values):
            row += f"{str(v):>4}"
            if i == 8:
                row += f"{'':>4}"
        print(row)
        print()

    format_row('#', holes)
    format_row('P', pars)
    format_row('H', hcps)
    for player in player_names:
        scores = [df.loc[df['hole'] == h, player].values[0] if player in df.columns else '' for h in holes]
        format_row(player, scores)

def compare_json_dataframes(json1: str, json2: str):
    """
    Loads two DataFrames from JSON files and compares them cell by cell, printing differing cells and the hole number if present.
    Args:
        json1 (str): Path to first JSON file.
        json2 (str): Path to second JSON file.
    """
    df1 = pd.DataFrame(json.load(open(json1, "r", encoding="utf-8")))
    df2 = pd.DataFrame(json.load(open(json2, "r", encoding="utf-8")))
    for row_idx in range(len(df1)):
        for col in df1.columns:
            val1 = df1.iloc[row_idx][col]
            val2 = df2.iloc[row_idx][col]
            if val1 != val2:
                hole = df1.iloc[row_idx]['hole'] if 'hole' in df1.columns else row_idx
                print(f"Cell ({row_idx}, '{col}') differs on hole {hole}: df1={val1} | df2={val2}")

def write_scorecard_to_excel(excel_path: str, json_path: str = "golf_df.json"):
    """
    Loads the scorecard DataFrame from a JSON file and writes par, hcp, and player scores to the given Excel file.
    Args:
        excel_path (str): Path to the Excel file.
        json_path (str): Path to the JSON file containing the DataFrame (default: 'golf_df.json').
    """
    from openpyxl import load_workbook
    import pandas as pd
    import json

    # Load DataFrame from JSON
    data = json.load(open(json_path, "r", encoding="utf-8"))
    df = pd.DataFrame(data)

    wb = load_workbook(excel_path)
    ws = wb["Scorecard"]

    # Write holes 1–9 to row 8, columns 4–12
    for i, hole in enumerate(range(1, 10), start=5):
        par_value = df.loc(df["hole"] == hole, "par").values[0]
        ws.cell(row=8, column=i, value=par_value)
        hcp_value = df.loc(df["hole"] == hole, "hcp").values[0]
        ws.cell(row=9, column=i, value=hcp_value)

    # Write holes 10–18 to row 8, columns 14–22
    for j, hole in enumerate(range(10, 19), start=15):
        par_value = df.loc(df["hole"] == hole, "par").values[0]
        ws.cell(row=8, column=j, value=par_value)
        hcp_value = df.loc(df["hole"] == hole, "hcp").values[0]
        ws.cell(row=9, column=j, value=hcp_value)

    # Write player scores
    player_names = [col for col in df.columns if col not in ["hole", "par", "hcp"]]
    for player in player_names:
        # Find the row for the player name in column 1
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip().lower() == player.lower():
                # Write scores for holes 1-9 (columns 5-13)
                for i, hole in enumerate(range(1, 10), start=5):
                    score = df.loc(df["hole"] == hole, player).values[0] if player in df.columns else None
                    ws.cell(row=row, column=i, value=score)
                # Write scores for holes 10-18 (columns 15-23)
                for j, hole in enumerate(range(10, 19), start=15):
                    score = df.loc(df["hole"] == hole, player).values[0] if player in df.columns else None
                    ws.cell(row=row, column=j, value=score)
                break  # Stop after finding the player

    wb.save(excel_path)
    print(f"Par values and player scores written to excel successfully! {excel_path}")

def clean_json_arrays(json_path):
    """
    Reads a JSON file, removes all spaces and tabs from all string elements in arrays, and writes it back.
    """
    import json
    def clean_item(item):
        if isinstance(item, list):
            return [clean_item(x) for x in item]
        elif isinstance(item, dict):
            return {k: clean_item(v) for k, v in item.items()}
        elif isinstance(item, str):
            return item.replace(" ","").replace("\t","")
        else:
            return item
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    cleaned = clean_item(data)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=2)
    print(f"Cleaned and saved: {json_path}")

def split_json_by_player(json_path, out_subfolder="json/players"):
    """
    Splits a JSON file with date keys and 'Spieler' dict into multiple files, one per player.
    Each output file contains only the data for that player under each date.
    Output files are placed in the given subfolder (created if missing).
    """
    import json
    import os
    os.makedirs(out_subfolder, exist_ok=True)
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Collect all unique player names
    player_set = set()
    for date_key, round_obj in data.items():
        spieler = round_obj.get("Spieler", {})
        player_set.update(spieler.keys())
    # For each player, build a new dict
    for player in player_set:
        out_data = {}
        for date_key, round_obj in data.items():
            if "Spieler" in round_obj and player in round_obj["Spieler"]:
                # Copy round_obj, but only keep this player in 'Spieler'
                new_round = dict(round_obj)
                new_round["Spieler"] = {player: round_obj["Spieler"][player]}
                out_data[date_key] = new_round
        # Write to file in subfolder
        out_path = os.path.join(out_subfolder, f"{player}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(out_data, f, ensure_ascii=False, indent=2)
        print(f"Wrote {out_path}")

def merge_player_jsons(players_folder, output_json_path):
    """
    Merges all player JSON files in a folder into one JSON file, combining all player data under each date.
    Output is written to output_json_path (e.g., '../golf_df_merged.json').
    """
    import os
    import json
    merged = {}
    for fname in os.listdir(players_folder):
        if fname.endswith('.json'):
            with open(os.path.join(players_folder, fname), 'r', encoding='utf-8') as f:
                pdata = json.load(f)
            for date_key, round_obj in pdata.items():
                if date_key not in merged:
                    merged[date_key] = dict(round_obj)
                    merged[date_key]["Spieler"] = {}
                # Add/merge this player's data
                for player, pval in round_obj.get("Spieler", {}).items():
                    merged[date_key]["Spieler"][player] = pval
    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Merged player JSONs to {output_json_path}")

def merge_json_files_on_date(json_path1, json_path2):
    """
    Merges two JSON files at the top-level date key. Makes a backup copy of the first JSON, then merges all date keys from the second JSON into the first file (in-place), overwriting any overlapping keys.
    The backup is written to <first_file>_copy_<date_time>.json in the same folder as the first file.
    """
    import json
    import os
    from datetime import datetime
    # Read both files
    with open(json_path1, "r", encoding="utf-8") as f1:
        data1 = json.load(f1)
    with open(json_path2, "r", encoding="utf-8") as f2:
        data2 = json.load(f2)
    # Make a backup copy of the first file
    folder = os.path.dirname(json_path1)
    base = os.path.splitext(os.path.basename(json_path1))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(folder, f"{base}_copy_{timestamp}.json")
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(data1, f, ensure_ascii=False, indent=2)
    # Merge: overwrite or add all keys from the second file
    merged = dict(data1)
    for k, v in data2.items():
        merged[k] = v  # Overwrite or add
    # Overwrite the first file with the merged result
    with open(json_path1, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Backup of {json_path1} written to {backup_path}")
    print(f"Merged {json_path2} into {json_path1} (overwrote {len(set(data1.keys()) & set(data2.keys()))} duplicate date keys)")

def merge_golf_df_jsons(json_path1, json_path2, output_path="json/golf_df.json"):
    """
    Merges two golf_df-like JSON files into one. For each date key, copies all Spieler entries from the second file to the first file,
    overwriting if the Spieler is already present. Keeps all other data from the first file. Makes backup copies of both files with date and time.
    The merged file is written to output_path (default: json/golf_df.json).
    """
    import json
    import os
    from datetime import datetime
    # Backup both files
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    for src in [json_path1, json_path2]:
        folder = os.path.dirname(src)
        base = os.path.splitext(os.path.basename(src))[0]
        backup_path = os.path.join(folder, f"{base}_backup_{now}.json")
        with open(src, "r", encoding="utf-8") as f:
            data = json.load(f)
        with open(backup_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    # Load both files
    with open(json_path1, "r", encoding="utf-8") as f1:
        data1 = json.load(f1)
    with open(json_path2, "r", encoding="utf-8") as f2:
        data2 = json.load(f2)
    # Merge logic
    merged = json.loads(json.dumps(data1))  # deep copy
    for date_key, round_obj2 in data2.items():
        if date_key not in merged:
            merged[date_key] = round_obj2
        else:
            # Merge Spieler dicts
            spieler1 = merged[date_key].get("Spieler", {})
            spieler2 = round_obj2.get("Spieler", {})
            for player, pdata in spieler2.items():
                spieler1[player] = pdata  # Overwrite or add
            merged[date_key]["Spieler"] = spieler1
    # Write merged file
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Merged file written to {output_path}")

def merge_golf_df_jsons_in_folder(folder_path, output_path="json/golf_df.json"):
    """
    Merges all golf_df-like JSON files in a folder into one file. For each date key, copies all Spieler entries from each file to the merged file,
    overwriting if the Spieler is already present. Keeps all other data from the first file encountered.
    The merged file is written to output_path (default: json/golf_df.json).
    No backup copies are made.
    """
    import json
    import os
    # Find all .json files in the folder
    file_list = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.json')]
    if not file_list:
        print(f"No JSON files found in {folder_path}")
        return
    # Merge all files
    merged = {}
    for file_path in file_list:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for date_key, round_obj in data.items():
            if date_key not in merged:
                merged[date_key] = round_obj
            else:
                # Merge Spieler dicts
                spieler1 = merged[date_key].get("Spieler", {})
                spieler2 = round_obj.get("Spieler", {})
                for player, pdata in spieler2.items():
                    spieler1[player] = pdata  # Overwrite or add
                merged[date_key]["Spieler"] = spieler1
    # Write merged file
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Merged {len(file_list)} files to {output_path}")

def create_default_golf_json(json_path: str, datum: str, ort: str, spieler_list: list):
    """
    Creates a default golf_df-like JSON file with the given date, location, and player names.
    All values (including arrays for Par, Hcp, Score, NettoP) are set to None.
    Output is written to json_path.
    """
    import json
    # Build default round data
    round_data = {
        "Ort": ort,
        "Par": [None]*18,
        "Hcp": [None]*18,
        "Spieler": {}
    }
    for name in spieler_list:
        round_data["Spieler"][name] = {
            "Platz": None,
            "Netto": None,
            "Geld": None,
            "Gesp.Hcp": None,
            "Birdies": None,
            "Pars": None,
            "Bogies": None,
            "Strich": None,
            "DayHcp": None,
            "Ladies": None, 
            "N2TP": None,
            "LD": None,
            "Score": [None]*18,
            "NettoP": [None]*18
        }
    out = {datum: round_data}
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"Default golf JSON written to {json_path}")

def copy_scorecard_to_golf_df(source_json_path, target_json_path):
    """
    Copies 'Hole', 'Par', 'Hcp', and all player score arrays from source_json_path (Vision model output)
    to target_json_path (e.g., golf_df.json), making a timestamped backup of the target file before overwriting.
    Player scores are written to the 'Score' key inside the 'Spieler' dict under the correct date key in the target JSON.
    Backup file will have .json extension.
    """
    # Make backup of target file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{target_json_path}_backup_{timestamp}.json"
    shutil.copy2(target_json_path, backup_path)

    # Load source and target JSONs
    with open(source_json_path, "r", encoding="utf-8") as f:
        source = json.load(f)
    with open(target_json_path, "r", encoding="utf-8") as f:
        target = json.load(f)

    # Find the date key in target (use today's date if present, else the first key)
    today = datetime.now().strftime("%d.%m.%Y")
    date_key = today if today in target else next((k for k in target.keys() if k not in ["Hole", "Par", "Hcp"]), None)
    if not date_key:
        print("No valid date key found in target JSON.")
        return

    # Copy keys
    for key in ["Hole", "Par", "Hcp"]:
        if key in source:
            target[date_key][key] = source[key]
    # Copy all player score arrays to 'Score' in 'Spieler' under date_key
    for player in [k for k in source.keys() if k not in ["Hole", "Par", "Hcp"]]:
        if "Spieler" in target[date_key] and player in target[date_key]["Spieler"]:
            target[date_key]["Spieler"][player]["Score"] = source[player]
        else:
            print(f"Warning: Player '{player}' not found in target[date_key]['Spieler'].")

    # Save updated target
    with open(target_json_path, "w", encoding="utf-8") as f:
        json.dump(target, f, ensure_ascii=False, indent=2)
    print(f"Scorecard data copied from {source_json_path} to {target_json_path}. Backup saved as {backup_path}.")

def calculate_money_for_players(json_path: str, date_key: str):
    """
    Calculates money for each player in golf_df.json at the given date.
    Money = Ladies * 10 + bonus for Platz:
    - Last (highest Platz): 40
    - Second last: 30
    - Third last: 20
    - Fourth last: 10
    - Rest: 0
    Updates the 'Geld' field for each player and saves the file.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if date_key not in data or "Spieler" not in data[date_key]:
        print(f"Date {date_key} not found or no players.")
        return
    players = data[date_key]["Spieler"]
    platz_list = [(pname, pdata.get("Platz", None)) for pname, pdata in players.items() if isinstance(pdata.get("Platz", None), int)]
    # Sort by Platz ascending (1 is best)
    platz_sorted = sorted(platz_list, key=lambda x: x[1])
    # Assign bonus
    n = len(platz_sorted)
    bonus_map = {}
    if n >= 1:
        bonus_map[platz_sorted[-1][0]] = 40
    if n >= 2:
        bonus_map[platz_sorted[-2][0]] = 30
    if n >= 3:
        bonus_map[platz_sorted[-3][0]] = 20
    if n >= 4:
        bonus_map[platz_sorted[-4][0]] = 10
    # Calculate money
    for pname, pdata in players.items():
        ladies = pdata.get("Ladies", 0) or 0
        bonus = bonus_map.get(pname, 0)
        pdata["Geld"] = int(ladies) * 10 + bonus
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Money calculated and saved for {date_key} in {json_path}")

def main():
    # Uncomment to run comparison
    # compare_json_dataframes("golf_df.json", "golf_df_real.json")
    print_scorecard_table_from_json("golf_df_real.json")
    write_scorecard_to_excel("2012_Mallorca_GRUPPO_SPORTIVO_STATISTIKO.xlsm", "golf_df.json")




if __name__ == "__main__":
    main()
